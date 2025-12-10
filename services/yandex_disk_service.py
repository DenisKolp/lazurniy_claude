"""
Yandex Disk integration service
"""
import json
import logging
import os
import io
from typing import Optional, List, Dict
from datetime import datetime, UTC
from zoneinfo import ZoneInfo
import aiohttp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class YandexDiskService:
    """Service for exporting voting results to Excel and uploading to Yandex Disk"""

    BASE_URL = "https://cloud-api.yandex.net/v1/disk"

    def __init__(self, oauth_token: Optional[str] = None, folder_path: Optional[str] = None):
        """
        Initialize Yandex Disk service

        Args:
            oauth_token: OAuth token for Yandex Disk API
            folder_path: Path to folder on Yandex Disk where files will be uploaded
        """
        # Try to get token from: parameter > env var > file
        self.oauth_token = oauth_token or os.getenv("YANDEX_DISK_TOKEN")

        # If not found, try to load from yandex_token.txt
        if not self.oauth_token and os.path.exists("yandex_token.txt"):
            try:
                with open("yandex_token.txt", "r") as f:
                    self.oauth_token = f.read().strip()
                logger.info("Loaded Yandex Disk token from yandex_token.txt")
            except Exception as e:
                logger.warning(f"Failed to load yandex_token.txt: {e}")

        self.folder_path = folder_path or os.getenv("YANDEX_DISK_FOLDER", "/Lazurny_Bot")
        self.session = None

        if self.oauth_token:
            logger.info(f"Yandex Disk service initialized (folder: {self.folder_path})")
        else:
            logger.warning("Yandex Disk service not available: no OAuth token")

    async def _get_session(self) -> aiohttp.ClientSession:
        """Get or create aiohttp session"""
        if self.session is None or self.session.closed:
            self.session = aiohttp.ClientSession(
                headers={
                    "Authorization": f"OAuth {self.oauth_token}",
                    "Content-Type": "application/json"
                }
            )
        return self.session

    async def close(self):
        """Close aiohttp session"""
        if self.session and not self.session.closed:
            await self.session.close()

    async def _create_folder(self, path: str) -> bool:
        """Create folder on Yandex Disk if it doesn't exist"""
        try:
            session = await self._get_session()
            url = f"{self.BASE_URL}/resources"
            params = {"path": path}

            async with session.put(url, params=params) as response:
                if response.status == 201:
                    logger.info(f"Created folder: {path}")
                    return True
                elif response.status == 409:
                    # Folder already exists
                    logger.debug(f"Folder already exists: {path}")
                    return True
                else:
                    text = await response.text()
                    logger.error(f"Failed to create folder: {response.status} - {text}")
                    return False
        except Exception as e:
            logger.error(f"Error creating folder: {e}")
            return False

    async def upload_file(self, file_content: bytes, file_name: str,
                         folder_path: Optional[str] = None) -> Optional[str]:
        """
        Upload file to Yandex Disk

        Args:
            file_content: File content as bytes
            file_name: Name of the file
            folder_path: Optional custom folder path (uses self.folder_path if not specified)

        Returns:
            Public URL to the file or None if failed
        """
        if not self.oauth_token:
            logger.warning("Yandex Disk not available: no OAuth token")
            return None

        try:
            # Ensure folder exists
            target_folder = folder_path or self.folder_path
            await self._create_folder(target_folder)

            # Get upload URL
            session = await self._get_session()
            file_path = f"{target_folder}/{file_name}"

            url = f"{self.BASE_URL}/resources/upload"
            params = {
                "path": file_path,
                "overwrite": "true"
            }

            async with session.get(url, params=params) as response:
                if response.status != 200:
                    text = await response.text()
                    logger.error(f"Failed to get upload URL: {response.status} - {text}")
                    return None

                data = await response.json()
                upload_url = data.get("href")

            if not upload_url:
                logger.error("No upload URL received")
                return None

            # Upload file
            async with session.put(upload_url, data=file_content) as response:
                if response.status not in (200, 201):
                    text = await response.text()
                    logger.error(f"Failed to upload file: {response.status} - {text}")
                    return None

            logger.info(f"File uploaded successfully: {file_path}")

            # Publish file to get public URL
            publish_url = f"{self.BASE_URL}/resources/publish"
            params = {"path": file_path}

            async with session.put(publish_url, params=params) as response:
                if response.status != 200:
                    text = await response.text()
                    logger.warning(f"Failed to publish file: {response.status} - {text}")
                    # Return disk URL even if publishing failed
                    return f"https://disk.yandex.ru/client/disk{file_path}"

            # Get public URL
            info_url = f"{self.BASE_URL}/resources"
            params = {"path": file_path}

            async with session.get(info_url, params=params) as response:
                if response.status != 200:
                    text = await response.text()
                    logger.warning(f"Failed to get file info: {response.status} - {text}")
                    return f"https://disk.yandex.ru/client/disk{file_path}"

                data = await response.json()
                public_url = data.get("public_url")

                if public_url:
                    logger.info(f"File published: {public_url}")
                    return public_url
                else:
                    return f"https://disk.yandex.ru/client/disk{file_path}"

        except Exception as e:
            logger.error(f"Error uploading file to Yandex Disk: {e}")
            return None

    def _create_voting_excel(
        self,
        voting_id: int,
        voting_title: str,
        voting_description: str,
        options: List[str],
        results: Dict[int, int],
        total_votes: int,
        created_at: datetime,
        ends_at: datetime
    ) -> bytes:
        """
        Create Excel file with voting results

        Returns:
            Excel file content as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты"

        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Голосование - Результаты"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="3399FF", end_color="3399FF", fill_type="solid")

        # Info section
        row = 3
        info_data = [
            ("Название:", voting_title),
            ("Описание:", voting_description),
            ("ID голосования:", str(voting_id)),
            ("Создано:", created_at.strftime("%d.%m.%Y %H:%M")),
            ("Завершено:", ends_at.strftime("%d.%m.%Y %H:%M")),
            ("Всего голосов:", str(total_votes))
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Results header
        row += 1
        header_row = row
        headers = ["Вариант ответа", "Количество голосов", "Процент"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Results data
        row += 1
        max_votes = max(results.values()) if results else 0

        for i, option in enumerate(options):
            votes = results.get(i, 0)
            percent = (votes / total_votes * 100) if total_votes > 0 else 0

            ws[f'A{row}'] = option
            ws[f'B{row}'] = votes
            ws[f'C{row}'] = f"{percent:.1f}%"

            # Highlight winning option
            if votes == max_votes and max_votes > 0:
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")

            row += 1

        # Footer
        row += 1
        moscow_tz = ZoneInfo('Europe/Moscow')
        export_time = datetime.now(UTC).astimezone(moscow_tz)
        ws[f'A{row}'] = f"Экспортировано: {export_time.strftime('%d.%m.%Y %H:%M MSK')}"

        # Auto-adjust column widths
        for col in range(1, 4):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def export_voting_results(
        self,
        voting_id: int,
        voting_title: str,
        voting_description: str,
        options: List[str],
        results: Dict[int, int],
        total_votes: int,
        created_at: datetime,
        ends_at: datetime,
        file_name: Optional[str] = None
    ) -> Optional[str]:
        """
        Export voting results to Excel and upload to Yandex Disk

        Returns:
            URL to the uploaded file or None if failed
        """
        if not self.oauth_token:
            logger.warning("Yandex Disk not available, skipping export")
            return None

        try:
            # Create Excel file
            excel_content = self._create_voting_excel(
                voting_id, voting_title, voting_description,
                options, results, total_votes, created_at, ends_at
            )

            # Generate file name
            safe_title = "".join(c for c in voting_title if c.isalnum() or c in (' ', '-', '_'))[:50]
            file_name = file_name or f"Голосование_{voting_id}_{safe_title}.xlsx"

            # Upload to Yandex Disk
            url = await self.upload_file(excel_content, file_name)

            if url:
                logger.info(f"Voting results exported to Yandex Disk: {url}")

            return url

        except Exception as e:
            logger.error(f"Failed to export voting results: {e}")
            return None

    def _create_registry_excel(self, members: List[Dict]) -> bytes:
        """
        Create Excel file with members registry

        Returns:
            Excel file content as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр"

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Реестр членов ассоциации КП 'Лазурный'"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="3399FF", end_color="3399FF", fill_type="solid")

        # Header
        row = 3
        headers = ["№", "ФИО", "Username", "Телефон", "Адрес участка", "Дата верификации"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Members data
        row = 4
        for idx, member in enumerate(members, 1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = member.get('full_name', 'Не указано')
            ws[f'C{row}'] = f"@{member.get('username', 'N/A')}"
            ws[f'D{row}'] = member.get('phone_number', 'Не указан')
            ws[f'E{row}'] = member.get('address', 'Не указан')
            ws[f'F{row}'] = member.get('verified_at', 'Не указана')

            # Zebra striping
            if idx % 2 == 0:
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            row += 1

        # Footer
        row += 1
        moscow_tz = ZoneInfo('Europe/Moscow')
        update_time = datetime.now(UTC).astimezone(moscow_tz)
        ws[f'A{row}'] = f"Всего членов: {len(members)}"
        ws[f'F{row}'] = f"Обновлено: {update_time.strftime('%d.%m.%Y %H:%M MSK')}"

        # Auto-adjust column widths
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def export_members_registry(
        self,
        members: List[Dict],
        file_name: Optional[str] = None
    ) -> Optional[str]:
        """
        Export members registry to Excel and upload to Yandex Disk

        Returns:
            URL to the uploaded file or None if failed
        """
        if not self.oauth_token:
            logger.warning("Yandex Disk not available, skipping export")
            return None

        try:
            # Create Excel file
            excel_content = self._create_registry_excel(members)

            # Generate file name
            file_name = file_name or "Реестр_членов_ассоциации_КП_Лазурный.xlsx"

            # Upload to Yandex Disk
            url = await self.upload_file(excel_content, file_name)

            if url:
                logger.info(f"Members registry exported to Yandex Disk: {url}")

            return url

        except Exception as e:
            logger.error(f"Failed to export members registry: {e}")
            return None

    async def export_all_voting_results(
        self,
        voting_results_list: List[Dict],
        file_name: Optional[str] = None
    ) -> Optional[str]:
        """
        Export all voting results to a single Excel file with multiple sheets

        Args:
            voting_results_list: List of dictionaries with voting data
            file_name: Optional custom file name

        Returns:
            URL to the uploaded file or None if failed
        """
        if not self.oauth_token:
            logger.warning("Yandex Disk not available, skipping export")
            return None

        if not voting_results_list:
            logger.warning("No voting results to export")
            return None

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create a worksheet for each voting question
            for idx, result_data in enumerate(voting_results_list):
                voting = result_data['voting']
                options = result_data['options']
                results = result_data['results']
                total_votes = result_data['total_votes']

                ws = wb.create_sheet(title=f"Вопрос {idx + 1}")

                # Title
                ws.merge_cells('A1:C1')
                title_cell = ws['A1']
                title_cell.value = f"Вопрос {idx + 1} - Результаты"
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
                title_cell.fill = PatternFill(start_color="3399FF", end_color="3399FF", fill_type="solid")

                # Info section
                row = 3
                info_data = [
                    ("Название:", voting.title),
                    ("Описание:", voting.description),
                    ("ID голосования:", str(voting.id)),
                    ("Создано:", voting.created_at.strftime("%d.%m.%Y %H:%M")),
                    ("Завершено:", voting.ends_at.strftime("%d.%m.%Y %H:%M")),
                    ("Всего голосов:", str(total_votes))
                ]

                for label, value in info_data:
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = value
                    row += 1

                # Results header
                row += 1
                headers = ["Вариант ответа", "Количество голосов", "Процент"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                # Results data
                row += 1
                max_votes = max(results.values()) if results else 0

                for i, option in enumerate(options):
                    votes = results.get(i, 0)
                    percent = (votes / total_votes * 100) if total_votes > 0 else 0

                    ws[f'A{row}'] = option
                    ws[f'B{row}'] = votes
                    ws[f'C{row}'] = f"{percent:.1f}%"

                    # Highlight winning option
                    if votes == max_votes and max_votes > 0:
                        for col in range(1, 4):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")

                    row += 1

                # Footer
                row += 1
                moscow_tz = ZoneInfo('Europe/Moscow')
                export_time = datetime.now(UTC).astimezone(moscow_tz)
                ws[f'A{row}'] = f"Экспортировано: {export_time.strftime('%d.%m.%Y %H:%M MSK')}"

                # Auto-adjust column widths
                for col in range(1, 4):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for cell in ws[column_letter]:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_content = output.getvalue()

            # Generate file name
            today = datetime.now(UTC).strftime('%d.%m.%Y')
            file_name = file_name or f"Результаты_голосования_{today}.xlsx"

            # Upload to Yandex Disk
            url = await self.upload_file(excel_content, file_name)

            if url:
                logger.info(f"All voting results exported to Yandex Disk: {url}")

            return url

        except Exception as e:
            logger.error(f"Failed to export all voting results: {e}")
            return None


# Global instance
yandex_disk_service = YandexDiskService()
